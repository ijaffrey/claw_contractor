"""Export takeoff data to Excel using openpyxl.

One sheet per trade, totals row, OH+profit row, grand total.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=12)
CURRENCY_FMT = '#,##0'
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)


def export_takeoff_to_excel(takeoff: dict[str, Any], out_path: Path) -> Path:
    """Write the takeoff to an .xlsx file at *out_path*."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    grand_subtotal = 0.0

    for trade in takeoff.get("trades", []):
        name = (trade.get("name") or "Trade")[:31]  # sheet name max 31 chars
        ws = wb.create_sheet(title=name.title())

        # Header row
        headers = ["Item", "Qty", "Unit", "Unit Rate", "Subtotal", "Confidence"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Line items
        row = 2
        trade_total = 0.0
        for li in trade.get("line_items", []):
            ws.cell(row=row, column=1, value=li.get("item", ""))
            ws.cell(row=row, column=2, value=li.get("quantity", 0))
            ws.cell(row=row, column=3, value=li.get("unit", ""))
            rate = li.get("unit_rate", 0) or 0
            ws.cell(row=row, column=4, value=rate).number_format = CURRENCY_FMT
            sub = li.get("subtotal", 0) or 0
            ws.cell(row=row, column=5, value=sub).number_format = CURRENCY_FMT
            ws.cell(row=row, column=6, value=li.get("confidence", ""))
            trade_total += sub
            for c in range(1, 7):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        # Totals row
        row += 1
        pricing = trade.get("pricing", {})
        subtotal = pricing.get("subtotal", trade_total)
        grand_subtotal += pricing.get("total", subtotal)

        ws.cell(row=row, column=4, value="Subtotal").font = TOTAL_FONT
        ws.cell(row=row, column=5, value=subtotal).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5).font = TOTAL_FONT
        row += 1

        if pricing.get("overhead"):
            ws.cell(row=row, column=4, value=f"Overhead ({(pricing.get('overhead_pct',0)*100):.0f}%)")
            ws.cell(row=row, column=5, value=pricing["overhead"]).number_format = CURRENCY_FMT
            row += 1
        if pricing.get("profit"):
            ws.cell(row=row, column=4, value=f"Profit ({(pricing.get('profit_pct',0)*100):.0f}%)")
            ws.cell(row=row, column=5, value=pricing["profit"]).number_format = CURRENCY_FMT
            row += 1
        if pricing.get("contingency"):
            ws.cell(row=row, column=4, value=f"Contingency ({(pricing.get('contingency_pct',0)*100):.0f}%)")
            ws.cell(row=row, column=5, value=pricing["contingency"]).number_format = CURRENCY_FMT
            row += 1

        total = pricing.get("total", subtotal)
        ws.cell(row=row, column=4, value="Total").font = Font(bold=True, size=13)
        ws.cell(row=row, column=5, value=total).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5).font = Font(bold=True, size=13)

        # Column widths
        widths = [30, 10, 8, 14, 14, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws = wb.create_sheet(title="Summary", index=0)
    ws.cell(row=1, column=1, value="Sanz Construction — Proposal Summary").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=takeoff.get("input_address", ""))
    ws.cell(row=4, column=1, value="Trade").font = HEADER_FONT
    ws.cell(row=4, column=2, value="Total").font = HEADER_FONT
    row = 5
    for trade in takeoff.get("trades", []):
        ws.cell(row=row, column=1, value=(trade.get("name") or "").title())
        total = trade.get("pricing", {}).get("total", 0)
        ws.cell(row=row, column=2, value=total).number_format = CURRENCY_FMT
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="Grand Total").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=grand_subtotal).number_format = CURRENCY_FMT
    ws.cell(row=row, column=2).font = TOTAL_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16

    wb.save(str(out_path))
    return out_path
